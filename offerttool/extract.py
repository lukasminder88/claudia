"""Schritt EXTRACT und PARSE: Zellen -> Rohwerte (Abschnitt 4 und 6).

Das Ergebnis ist ein ``StandortContext`` je Kalktool.  Der Schritt ist
seiteneffektfrei und berührt das Dokument nicht.
"""

from __future__ import annotations

import datetime as _dt
import re
from dataclasses import dataclass, field
from decimal import Decimal

from . import parsers
from .errors import WarningCollector
from .formatters import _to_decimal, trim
from .mapping import ListSpec, Mapping
from .workbook import Kalktool


def _num(value) -> Decimal:
    """Zahlwert einer Zelle; Text und Fehlerwerte ergeben 0."""
    if isinstance(value, str) and value.strip().startswith("#"):
        return Decimal(0)
    return _to_decimal(value)


def _is_error(value) -> bool:
    return isinstance(value, str) and value.strip().startswith("#")


@dataclass
class Position:
    """Eine Zeile einer Positionsliste (Abschnitt 4.5)."""

    bezeichnung: str = ""
    artnr: str = "–"
    stueck: str = "1"
    betrag: Decimal = Decimal(0)
    extra: dict = field(default_factory=dict)


@dataclass
class StandortContext:
    """Alle Rohwerte eines Standorts, vor der Ableitung."""

    quelle: str
    index: int = 1
    values: dict = field(default_factory=dict)
    listen: dict[str, list[Position]] = field(default_factory=dict)
    probes: dict = field(default_factory=dict)
    warn: WarningCollector = field(default_factory=WarningCollector)
    blocked_values: list = field(default_factory=list)

    def get(self, name, default=None):
        return self.values.get(name, default)

    def num(self, name) -> Decimal:
        return _num(self.values.get(name))

    def text(self, name) -> str:
        return trim(self.values.get(name))


def extract(kalktool: Kalktool, mapping: Mapping, warn: WarningCollector) -> StandortContext:
    """Feldkatalog und Positionslisten eines Kalktools einlesen."""
    ctx = StandortContext(quelle=kalktool.path.name, warn=warn)

    for name, spec in mapping.fields.items():
        raw = kalktool.cell(spec.cell)
        ctx.values[name] = raw

    # Freitextfelder zerlegen (Abschnitt 6)
    ctx.values["kunde.plz_ort"] = parsers.parse_plz_ort(
        ctx.values.get("kunde.plz_ort_roh"), warn, "kunde"
    )
    ctx.values["standort.plz_ort"] = parsers.parse_plz_ort(
        ctx.values.get("standort.plz_ort_roh"), warn, "standort"
    )
    ctx.values["kunde.kontakt"] = parsers.parse_kontakt(ctx.values.get("kunde.kontakt_roh"), warn)
    _standort_dito(ctx, warn)
    ctx.values["vertragsbeginn"] = parsers.parse_vertragsbeginn(
        ctx.values.get("vertragsbeginn_roh"), warn
    )

    # Datumsfeld M9: der zwischengespeicherte Wert einer =TODAY()-Formel ist das
    # Öffnungsdatum, nicht das Offertdatum (Abschnitt 6.4).
    datum = ctx.values.get("datum")
    if isinstance(datum, _dt.datetime):
        ctx.values["datum"] = datum.date()
    if _formula_is_today(kalktool, mapping):
        warn.add("W308", "KM!M9 enthält =TODAY()")

    for name, ref in mapping.probes.items():
        ctx.probes[name] = kalktool.cell(ref, allow_blocked=True)
    if _is_error(ctx.probes.get("divzero.hw")) or _is_error(ctx.probes.get("divzero.sw")):
        warn.add("W312", "Rabattsatz ohne Listenpreis")

    for name, spec in mapping.lists.items():
        ctx.listen[name] = _read_list(kalktool, spec, warn)

    _check_stueck(ctx, kalktool, mapping, warn)
    ctx.blocked_values = kalktool.blocked_values()
    return ctx


# Kürzel, mit denen im Kalktool auf den Kunden verwiesen wird, statt den
# Standort noch einmal auszuschreiben.
# Punkte und Leerzeichen sind hier bedeutungslos: "s. o." und "s.o." meinen
# dasselbe.
DITO = {"dito", "ditto", "dto", "idem", "so", "sieheoben", "wieoben", "gleich"}


def _standort_dito(ctx: StandortContext, warn: WarningCollector) -> None:
    """„dito“ als Standortname durch den Kundennamen ersetzen."""
    name = trim(ctx.values.get("standort.name"))
    if re.sub(r"[.\s]+", "", name).lower() not in DITO:
        return
    kunde = trim(ctx.values.get("kunde.firma"))
    if not kunde:
        return
    ctx.values["standort.name"] = kunde
    warn.add("W317", f"{name!r} -> {kunde!r}")


def _formula_is_today(kalktool: Kalktool, mapping: Mapping) -> bool:
    """Prüfen, ob die Datumszelle eine ``TODAY()``-Formel trägt."""
    import warnings as _w

    import openpyxl

    spec = mapping.fields.get("datum")
    if spec is None:
        return False
    try:
        with _w.catch_warnings():
            _w.simplefilter("ignore")
            wb = openpyxl.load_workbook(kalktool.path, data_only=False, read_only=True)
        try:
            ws = wb.worksheets[mapping.sheets[spec.cell.sheet]["index"]]
            raw = ws.cell(row=spec.cell.row, column=spec.cell.col).value
        finally:
            wb.close()
    except Exception:  # pragma: no cover - defensiv
        return False
    return isinstance(raw, str) and "TODAY()" in raw.upper()


def _read_list(kalktool: Kalktool, spec: ListSpec, warn: WarningCollector) -> list[Position]:
    rows = kalktool.range_values(spec.range)
    if spec.layout == "paired":
        return _read_paired(rows, spec)
    return _read_rows(rows, spec)


def _cell(row: list, spec: ListSpec, colname: str):
    idx = spec.cols.get(colname)
    if idx is None:
        return None
    off = idx - spec.range.col1
    return row[off] if 0 <= off < len(row) else None


def _read_rows(rows: list[list], spec: ListSpec) -> list[Position]:
    out: list[Position] = []
    for row in rows:
        bez = trim(_cell(row, spec, "bezeichnung"))
        if not bez or bez == "Support :":
            continue
        betrag_col = "netto" if "netto" in spec.cols else ("total" if "total" in spec.cols else "listenpreis")
        betrag = _num(_cell(row, spec, betrag_col))
        if "netto <= 0" in spec.skip_if or "total <= 0" in spec.skip_if:
            if betrag <= 0:
                continue
        anzahl = _cell(row, spec, "anzahl") or _cell(row, spec, "stunden")
        artnr = trim(_cell(row, spec, "artnr")) or "–"
        stueck = trim(anzahl) if anzahl not in (None, "", 0) else (
            str(spec.stueck_fix) if spec.stueck_fix else ""
        )
        extra = {}
        if "preis_std" in spec.cols:
            extra["preis_std"] = _num(_cell(row, spec, "preis_std"))
        if "listenpreis" in spec.cols:
            extra["listenpreis"] = _num(_cell(row, spec, "listenpreis"))
        out.append(Position(bezeichnung=bez, artnr=artnr, stueck=stueck, betrag=betrag, extra=extra))
    return out


def _read_paired(rows: list[list], spec: ListSpec) -> list[Position]:
    """Blockstruktur ``dienstleistung``: Zeile n = Label, Zeile n+1 = Beträge."""
    from .formatters import label_clean

    out: list[Position] = []
    i = 0
    while i < len(rows) - 1:
        label_raw = _cell(rows[i], spec, spec.use_col or "verrechnet")
        label = trim(label_raw)
        # Eine Labelzeile ist eine Zeile mit Text in der genutzten Spalte.
        if not label or _num(label_raw) != 0 or isinstance(label_raw, (int, float)):
            i += 1
            continue
        wert_row = rows[i + 1]
        betrag = _num(_cell(wert_row, spec, spec.use_col or "verrechnet"))
        if betrag > 0:
            clean = spec.label_override.get(label)
            if clean is None:
                clean = label_clean(label_raw) if spec.label_fmt == "label_clean" else label
            out.append(Position(bezeichnung=clean, betrag=betrag))
        i += 2
    return out


def _check_stueck(ctx: StandortContext, kalktool: Kalktool, mapping: Mapping, warn) -> None:
    """``KM!C53 == sum(KM!C27:C52)`` plausibilisieren (Abschnitt 5.5).

    Der Wert von ``C53`` darf gelesen, aber nicht ausgegeben werden.
    """
    spec = mapping.lists.get("hardware")
    if spec is None or not spec.stueck_check:
        return
    total_ref = mapping.probes.get("stueck_summe")
    if total_ref is None:
        return
    total = _num(kalktool.cell(total_ref, allow_blocked=True))
    summe = sum(
        (p.extra.get("listenpreis", Decimal(0)) for p in ctx.listen.get("hardware", [])),
        Decimal(0),
    )
    ctx.probes["stueck_belegbar"] = abs(total - summe) <= Decimal("0.01")
    if not ctx.probes["stueck_belegbar"]:
        warn.add("W307", "Stückzahl nicht belegbar")
