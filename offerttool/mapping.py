"""Laden und Auswerten von ``mapping.yaml`` (Spezifikation V3, Abschnitt 15).

Der Generator enthält keine Zelladressen.  Eine neue Kalktool-Version wird durch
eine weitere YAML abgebildet; die Auswahl erfolgt über die Versionszelle
(``KM!C1``).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import yaml

from openpyxl.utils import column_index_from_string, get_column_letter

RESOURCES = Path(__file__).with_name("resources")

RE_CELL = re.compile(r"^(?P<sheet>[A-Za-z_][\w.]*)!(?P<col>[A-Z]{1,3})(?P<row>\d+)$")
RE_RANGE = re.compile(
    r"^(?P<sheet>[A-Za-z_][\w.]*)!(?P<c1>[A-Z]{1,3})(?P<r1>\d+):(?P<c2>[A-Z]{1,3})(?P<r2>\d+)$"
)


@dataclass(frozen=True)
class CellRef:
    sheet: str
    row: int
    col: int

    @property
    def a1(self) -> str:
        return f"{self.sheet}!{get_column_letter(self.col)}{self.row}"


@dataclass(frozen=True)
class RangeRef:
    sheet: str
    row1: int
    col1: int
    row2: int
    col2: int

    def contains(self, ref: CellRef) -> bool:
        return (
            ref.sheet == self.sheet
            and self.row1 <= ref.row <= self.row2
            and self.col1 <= ref.col <= self.col2
        )

    @property
    def a1(self) -> str:
        return (
            f"{self.sheet}!{get_column_letter(self.col1)}{self.row1}"
            f":{get_column_letter(self.col2)}{self.row2}"
        )


def parse_cell(text: str) -> CellRef:
    m = RE_CELL.match(text.strip())
    if not m:
        raise ValueError(f"Ungültige Zelladresse: {text!r}")
    return CellRef(m.group("sheet"), int(m.group("row")), column_index_from_string(m.group("col")))


def parse_range(text: str) -> RangeRef:
    m = RE_RANGE.match(text.strip())
    if not m:
        raise ValueError(f"Ungültiger Bereich: {text!r}")
    return RangeRef(
        m.group("sheet"),
        int(m.group("r1")),
        column_index_from_string(m.group("c1")),
        int(m.group("r2")),
        column_index_from_string(m.group("c2")),
    )


def parse_ref(text: str):
    """Adresse oder Bereich, je nach Schreibweise."""
    return parse_range(text) if ":" in text else parse_cell(text)


@dataclass
class FieldSpec:
    name: str
    cell: CellRef
    type: str
    fmt: str | None
    req: str
    values: list | None = None
    only: list[str] | None = None

    def required_for(self, variante: str | None) -> bool:
        if self.req != "M":
            return False
        if self.only and variante is not None:
            return variante in self.only
        return True


@dataclass
class ListSpec:
    name: str
    range: RangeRef
    cols: dict[str, int]
    skip_if: list[str]
    layout: str = "rows"
    use_col: str | None = None
    label_fmt: str | None = None
    label_override: dict[str, str] | None = None
    stueck_fix: int | None = None
    stueck_check: str | None = None


class Mapping:
    """Ein geladenes Mapping mit typisierten Zugriffshelfern."""

    def __init__(self, data: dict, source: Path | None = None) -> None:
        self.raw = data
        self.source = source
        self.version: str = data["version"]
        self.version_cell = parse_cell(data["version_cell"])
        self.sheets: dict[str, dict] = data["sheets"]
        self.styles: dict[str, str] = data.get("styles", {})
        self.probes = {k: parse_cell(v) for k, v in data.get("probes", {}).items()}
        self.layout_marken = {
            parse_cell(k): str(v) for k, v in (data.get("layout_marken") or {}).items()
        }

        self.fields: dict[str, FieldSpec] = {}
        for name, spec in data["fields"].items():
            self.fields[name] = FieldSpec(
                name=name,
                cell=parse_cell(spec["cell"]),
                type=spec["type"],
                fmt=spec.get("fmt"),
                req=spec.get("req", "O"),
                values=spec.get("values"),
                only=spec.get("only"),
            )

        self.lists: dict[str, ListSpec] = {}
        for name, spec in data["lists"].items():
            self.lists[name] = ListSpec(
                name=name,
                range=parse_range(spec["range"]),
                cols={k: column_index_from_string(v) for k, v in spec["cols"].items()},
                skip_if=list(spec.get("skip_if", [])),
                layout=spec.get("layout", "rows"),
                use_col=spec.get("use_col"),
                label_fmt=spec.get("label_fmt"),
                label_override=spec.get("label_override") or {},
                stueck_fix=spec.get("stueck_fix"),
                stueck_check=spec.get("stueck_check"),
            )

        self.blocked = [parse_ref(x) for x in data.get("blocked", [])]
        self.blocked_readable = {parse_cell(x).a1 for x in data.get("blocked_readable", [])}

    # -- Sperrliste ---------------------------------------------------------

    def is_blocked(self, ref: CellRef) -> bool:
        """Wahr, wenn die Zelle in der Sperrliste liegt (Abschnitt 13.2)."""
        for b in self.blocked:
            if isinstance(b, CellRef):
                if b == ref:
                    return True
            elif b.contains(ref):
                return True
        return False

    def is_readable_despite_block(self, ref: CellRef) -> bool:
        return ref.a1 in self.blocked_readable

    def sheet_key(self, name: str) -> str:
        return name

    def style(self, key: str, default: str = "Normal") -> str:
        return self.styles.get(key, default)


def load_mapping(path: str | Path) -> Mapping:
    p = Path(path)
    with p.open(encoding="utf-8") as fh:
        return Mapping(yaml.safe_load(fh), p)


def available_mappings() -> dict[str, Path]:
    """Alle mitgelieferten Mappings, nach Versionsstring."""
    out: dict[str, Path] = {}
    for p in sorted(RESOURCES.glob("mapping_*.yaml")):
        with p.open(encoding="utf-8") as fh:
            out[yaml.safe_load(fh)["version"]] = p
    return out


def _vergleichbar(text) -> str:
    """Beschriftungen vergleichbar machen: Leerraum und Gross-/Kleinschreibung."""
    import re as _re

    return _re.sub(r"\s+", " ", str(text or "")).strip().lower()


def layout_passt(mapping: Mapping, lies_zelle) -> tuple[bool, list[str]]:
    """Prüft die Beschriftungen des Kalktools gegen die Layoutmarken.

    ``lies_zelle`` bekommt eine :class:`CellRef` und liefert den Zellwert.
    Zurück kommt, ob alle Marken stimmen, und welche abweichen.
    """
    abweichungen = []
    for ref, erwartet in mapping.layout_marken.items():
        try:
            ist = lies_zelle(ref)
        except Exception:
            ist = None
        if _vergleichbar(ist) != _vergleichbar(erwartet):
            abweichungen.append(f"{ref.a1}: erwartet {erwartet!r}, gefunden {ist!r}")
    return (not abweichungen), abweichungen


def select_mapping(version_text: str) -> Mapping:
    """Mapping anhand des Inhalts der Versionszelle wählen.

    ``"Version: Q4 2025"`` wählt das Mapping mit ``version: "Q4 2025"``.
    """
    version_text = (version_text or "").strip()
    candidates = available_mappings()
    for version, path in candidates.items():
        if version and version in version_text:
            return load_mapping(path)
    raise KeyError(
        f"Kein Mapping für {version_text!r}; verfügbar: {sorted(candidates)}"
    )


def waehle_mapping(version_text: str, lies_zelle, warn) -> Mapping:
    """Mapping über die Versionszelle wählen, sonst über das Layout.

    Trägt ein Kalktool eine unbekannte Version, wird nicht geraten: die
    Beschriftungen werden gegen die Layoutmarken jedes bekannten Mappings
    gehalten. Passt genau eines, wird es mit Warnung ``W313`` verwendet;
    passt keines oder mehrere, bricht es mit ``E202`` ab.

    Auch bei bekannter Version werden die Marken geprüft – ein umgebautes
    Kalktool fällt so auf, statt still falsche Zellen zu liefern (``W314``).
    """
    from .errors import OfferteError

    version_text = (version_text or "").strip()
    verfuegbar = available_mappings()

    for version, pfad in verfuegbar.items():
        if version and version in version_text:
            mapping = load_mapping(pfad)
            passt, abweichungen = layout_passt(mapping, lies_zelle)
            if not passt:
                warn.add("W314", "; ".join(abweichungen[:3]))
            return mapping

    passende = []
    for version, pfad in verfuegbar.items():
        mapping = load_mapping(pfad)
        passt, _ = layout_passt(mapping, lies_zelle)
        if passt:
            passende.append((version, mapping))

    if len(passende) == 1:
        version, mapping = passende[0]
        warn.add("W313", f"{version_text or 'ohne Angabe'} → Mapping {version}")
        return mapping

    if len(passende) > 1:
        raise OfferteError(
            "E202",
            f"{version_text or 'ohne Versionsangabe'}: mehrere Mappings passen "
            f"({', '.join(v for v, _ in passende)}). Bitte mit --mapping wählen.",
        )
    raise OfferteError(
        "E202",
        f"{version_text or 'ohne Versionsangabe'}: das Layout stimmt mit keinem "
        f"bekannten Mapping überein (bekannt: {', '.join(sorted(verfuegbar))}). "
        "Entweder ist das Kalktool umgebaut, oder es braucht ein eigenes Mapping.",
    )
