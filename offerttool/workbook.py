"""Zugriff auf ein Kalktool (.xlsx).

Blätter werden **nach Position** gelesen (Abschnitt 2.2); der Blattname wird nur
als Warnung geprüft, damit ein umbenanntes Blatt die Generierung nicht stoppt.
Zellen der Sperrliste (Abschnitt 13.2) werden gar nicht erst herausgegeben.
"""

from __future__ import annotations

import datetime as _dt
import warnings
from pathlib import Path

import openpyxl

from .errors import OfferteError, WarningCollector
from .mapping import CellRef, Mapping, RangeRef, parse_cell


class Kalktool:
    """Ein geöffnetes Kalktool mit mapping-gebundenem Zellzugriff."""

    def __init__(self, path: str | Path, mapping: Mapping, warn: WarningCollector) -> None:
        self.path = Path(path)
        self.mapping = mapping
        self.warn = warn
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                self._wb = openpyxl.load_workbook(self.path, data_only=True, read_only=False)
        except Exception as exc:  # pragma: no cover - Dateisystem/Formatfehler
            raise OfferteError("E201", f"{self.path.name}: {exc}") from exc

        if len(self._wb.worksheets) < 2:
            raise OfferteError("E201", f"{self.path.name}: nur {len(self._wb.worksheets)} Blatt/Blätter")

        self._sheets = {}
        for key, spec in mapping.sheets.items():
            idx = spec["index"]
            if idx >= len(self._wb.worksheets):
                raise OfferteError("E201", f"{self.path.name}: Blattindex {idx} fehlt")
            ws = self._wb.worksheets[idx]
            expect = spec.get("expect_name")
            if expect is not None and ws.title != expect:
                warn.add("W320", f"{key}: erwartet {expect!r}, gefunden {ws.title!r}")
            self._sheets[key] = ws

    # -- Rohzugriff ---------------------------------------------------------

    def sheet(self, key: str):
        try:
            return self._sheets[key]
        except KeyError:
            raise OfferteError("E211", f"Unbekanntes Blatt {key!r}") from None

    def _raw(self, ref: CellRef):
        ws = self.sheet(ref.sheet)
        if ref.row > max(ws.max_row, 1) or ref.col > max(ws.max_column, 1):
            raise OfferteError("E211", f"{self.path.name}: {ref.a1} ausserhalb {ws.dimensions}")
        return ws.cell(row=ref.row, column=ref.col).value

    def cell(self, ref: CellRef | str, *, allow_blocked: bool = False):
        """Zellwert lesen.  Gesperrte Zellen liefern nur mit ``allow_blocked``."""
        if isinstance(ref, str):
            ref = parse_cell(ref)
        if self.mapping.is_blocked(ref) and not (
            allow_blocked and self.mapping.is_readable_despite_block(ref)
        ):
            raise OfferteError("E211", f"{ref.a1} steht auf der Sperrliste und wird nicht gelesen")
        return self._raw(ref)

    def range_values(self, rng: RangeRef) -> list[list]:
        """Rechteck als Zeilenliste; Zeilenindex 0 entspricht ``rng.row1``."""
        ws = self.sheet(rng.sheet)
        rows = []
        for r in range(rng.row1, rng.row2 + 1):
            if r > max(ws.max_row, 1):
                rows.append([None] * (rng.col2 - rng.col1 + 1))
                continue
            rows.append([ws.cell(row=r, column=c).value for c in range(rng.col1, rng.col2 + 1)])
        return rows

    def blocked_values(self) -> list:
        """Alle Werte der Sperrliste – Grundlage der Ausgabeprüfung E601."""
        out = []
        for b in self.mapping.blocked:
            refs = [b] if isinstance(b, CellRef) else [
                CellRef(b.sheet, r, c)
                for r in range(b.row1, b.row2 + 1)
                for c in range(b.col1, b.col2 + 1)
            ]
            for ref in refs:
                ws = self._sheets.get(ref.sheet)
                if ws is None or ref.row > max(ws.max_row, 1) or ref.col > max(ws.max_column, 1):
                    continue
                v = ws.cell(row=ref.row, column=ref.col).value
                if v is not None and not isinstance(v, (_dt.date, _dt.datetime, bool)):
                    out.append((ref.a1, v))
        return out

    def close(self) -> None:
        self._wb.close()

    def __enter__(self) -> "Kalktool":
        return self

    def __exit__(self, *exc) -> None:
        self.close()


def read_version_cell(path: str | Path, address: str = "KM!C1") -> str:
    """Versionszelle lesen, bevor ein Mapping gewählt ist."""
    ref = parse_cell(address)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(Path(path), data_only=True, read_only=True)
    try:
        if not wb.worksheets:
            raise OfferteError("E201", f"{Path(path).name}: keine Blätter")
        ws = wb.worksheets[0]
        return str(ws.cell(row=ref.row, column=ref.col).value or "")
    finally:
        wb.close()
